#!/usr/bin/env python3
"""Run Book2.xlsx rows through the XLSX calculation engine and write results back."""

from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment
from sqlalchemy import select

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
DEFAULT_INPUT = Path("/Users/rohitsundaram/Downloads/Book2.xlsx")
DEFAULT_OUTPUT = Path("/Users/rohitsundaram/Downloads/Book2_with_results.xlsx")
DEFAULT_REPORT = ROOT / "docs" / "BOOK2_BATCH_CALCULATION_REPORT.md"

CSV_FIELDNAMES = [
    "row_num",
    "trade",
    "client",
    "engineers",
    "hours",
    "parking",
    "congestion",
    "materials",
    "labour_charge",
    "materials_charge",
    "final_total",
    "profit_gbp",
    "profit_pct",
    "formula_source",
    "status",
    "error",
    "internal_notes",
]

OUTPUT_HEADERS = {
    8: "Labour charge to client",
    9: "Materials / Parking / CC charge",
    10: "Final total",
    11: "Profit GBP",
    12: "Profit %",
    13: "Formula source",
    14: "Status",
    15: "Error (if any)",
    16: "Internal notes",
}


@dataclass
class RowResult:
    row_num: int
    trade: str
    client: str
    status: str
    engineers: float | int | None = None
    hours: float | int | None = None
    parking: float | int | None = None
    congestion: float | int | None = None
    materials: float | int | None = None
    error: str | None = None
    labour_charge: float | None = None
    materials_charge: float | None = None
    final_total: float | None = None
    profit_gbp: float | None = None
    profit_pct: float | None = None
    formula_source: str | None = None
    internal_notes: str | None = None

    def to_csv_row(self) -> dict[str, object]:
        return {
            "row_num": self.row_num,
            "trade": self.trade,
            "client": self.client,
            "engineers": self.engineers,
            "hours": self.hours,
            "parking": self.parking,
            "congestion": self.congestion,
            "materials": self.materials,
            "labour_charge": self.labour_charge,
            "materials_charge": self.materials_charge,
            "final_total": self.final_total,
            "profit_gbp": self.profit_gbp,
            "profit_pct": self.profit_pct,
            "formula_source": self.formula_source,
            "status": self.status,
            "error": self.error,
            "internal_notes": self.internal_notes,
        }


def _to_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace("£", "")
    if not text:
        return Decimal("0")
    return Decimal(text)


def _setup_backend_path() -> None:
    if str(BACKEND) not in sys.path:
        sys.path.insert(0, str(BACKEND))


def _calculate_row(db, trade_name: str, client_name: str, engineers, hours, parking, congestion, materials):
    from app.core.config import settings
    from app.engines.approval_engine import build_calculation_breakdown
    from app.engines.rules_engine import find_active_rule
    from app.models.trade import Trade
    from app.schemas.calculation import ChargeInput, LabourInput, MaterialInput
    from app.services.client_service import find_client_by_name_or_alias

    client = find_client_by_name_or_alias(db, client_name)
    if client is None:
        raise ValueError(f"Client not found: {client_name}")

    trade = db.scalar(select(Trade).where(Trade.name == trade_name.strip()))
    if trade is None:
        raise ValueError(f"Trade not found: {trade_name}")

    matched = find_active_rule(db, client.id, trade.id, None)
    if matched is None:
        raise ValueError(f"No active rate rule for client '{client_name}' / trade '{trade_name}'")

    engineers_dec = _to_decimal(engineers)
    hours_dec = _to_decimal(hours)
    parking_dec = _to_decimal(parking)
    congestion_dec = _to_decimal(congestion)
    materials_dec = _to_decimal(materials)

    if engineers_dec <= 0:
        raise ValueError("Number of engineers must be greater than 0")
    if hours_dec <= 0:
        raise ValueError("Hours must be greater than 0")

    material_items = []
    if materials_dec > 0:
        material_items.append(
            MaterialInput(
                material_name="Materials",
                quantity=Decimal("1"),
                unit_cost=materials_dec,
            )
        )

    charges = None
    if parking_dec > 0 or congestion_dec > 0:
        charges = ChargeInput(
            parking_required=parking_dec > 0,
            parking_type="fixed" if parking_dec > 0 else None,
            parking_fixed_amount=parking_dec if parking_dec > 0 else None,
            congestion_required=congestion_dec > 0,
            congestion_amount=congestion_dec,
        )

    breakdown = build_calculation_breakdown(
        labour_items=[
            LabourInput(
                labour_type="hourly",
                number_of_engineers=int(engineers_dec),
                hours_on_site=hours_dec,
            )
        ],
        material_items=material_items,
        charges=charges,
        matched_rule=matched,
        formula_version=settings.formula_version,
    )

    return breakdown


def _calculate_days_row(db, trade_name: str, client_name: str, engineers, days, parking, congestion, materials):
    from app.core.config import settings
    from app.engines.approval_engine import build_calculation_breakdown
    from app.engines.rules_engine import find_active_rule
    from app.models.trade import Trade
    from app.schemas.calculation import ChargeInput, LabourInput, MaterialInput
    from app.services.client_service import find_client_by_name_or_alias

    client = find_client_by_name_or_alias(db, client_name)
    if client is None:
        raise ValueError(f"Client not found: {client_name}")

    trade = db.scalar(select(Trade).where(Trade.name == trade_name.strip()))
    if trade is None:
        raise ValueError(f"Trade not found: {trade_name}")

    matched = find_active_rule(db, client.id, trade.id, None)
    if matched is None:
        raise ValueError(f"No active rate rule for client '{client_name}' / trade '{trade_name}'")

    engineers_dec = _to_decimal(engineers)
    days_dec = _to_decimal(days)
    parking_dec = _to_decimal(parking)
    congestion_dec = _to_decimal(congestion)
    materials_dec = _to_decimal(materials)

    if engineers_dec <= 0:
        raise ValueError("Number of engineers must be greater than 0")
    if days_dec <= 0:
        raise ValueError("Days must be greater than 0")

    material_items = []
    if materials_dec > 0:
        material_items.append(
            MaterialInput(
                material_name="Materials",
                quantity=Decimal("1"),
                unit_cost=materials_dec,
            )
        )

    charges = None
    if parking_dec > 0 or congestion_dec > 0:
        charges = ChargeInput(
            parking_required=parking_dec > 0,
            parking_type="fixed" if parking_dec > 0 else None,
            parking_fixed_amount=parking_dec if parking_dec > 0 else None,
            congestion_required=congestion_dec > 0,
            congestion_amount=congestion_dec,
        )

    breakdown = build_calculation_breakdown(
        labour_items=[
            LabourInput(
                labour_type="day",
                number_of_engineers=int(engineers_dec),
                days_on_site=days_dec,
                number_of_labourers=0,
            )
        ],
        material_items=material_items,
        charges=charges,
        matched_rule=matched,
        formula_version=settings.formula_version,
    )

    return breakdown


def write_results_csv(results: list[RowResult], csv_path: Path) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_FIELDNAMES, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        for result in results:
            writer.writerow(result.to_csv_row())


def process_workbook(
    input_path: Path,
    output_path: Path | None,
    *,
    limit: int | None = None,
) -> list[RowResult]:
    _setup_backend_path()
    from app.db.session import SessionLocal

    wb_out = openpyxl.load_workbook(input_path) if output_path is not None else None
    ws = wb_out.active if wb_out is not None else None
    wb_data = openpyxl.load_workbook(input_path, data_only=True)
    ws_data = wb_data.active

    wrap = Alignment(wrap_text=True, vertical="top")
    if ws is not None:
        for col, header in OUTPUT_HEADERS.items():
            ws.cell(row=1, column=col, value=header)

        ws.column_dimensions["H"].width = 22
        ws.column_dimensions["I"].width = 28
        ws.column_dimensions["J"].width = 14
        ws.column_dimensions["K"].width = 12
        ws.column_dimensions["L"].width = 10
        ws.column_dimensions["M"].width = 14
        ws.column_dimensions["N"].width = 10
        ws.column_dimensions["O"].width = 40
        ws.column_dimensions["P"].width = 80
        ws.column_dimensions["P"].alignment = wrap

    results: list[RowResult] = []
    db = SessionLocal()
    processed = 0

    try:
        for row_num in range(2, ws_data.max_row + 1):
            trade = ws_data.cell(row=row_num, column=1).value
            if trade is None or str(trade).strip() == "":
                continue

            if limit is not None and processed >= limit:
                break

            client = ws_data.cell(row=row_num, column=2).value
            engineers = ws_data.cell(row=row_num, column=3).value
            hours = ws_data.cell(row=row_num, column=4).value
            parking = ws_data.cell(row=row_num, column=5).value
            congestion = ws_data.cell(row=row_num, column=6).value
            materials = ws_data.cell(row=row_num, column=7).value

            trade_str = str(trade).strip()
            client_str = str(client or "").strip()

            result = RowResult(
                row_num=row_num,
                trade=trade_str,
                client=client_str,
                status="ERROR",
                engineers=engineers,
                hours=hours,
                parking=parking,
                congestion=congestion,
                materials=materials,
            )

            try:
                breakdown = _calculate_row(
                    db,
                    trade_str,
                    client_str,
                    engineers,
                    hours,
                    parking,
                    congestion,
                    materials,
                )
                result.status = "OK"
                result.labour_charge = float(breakdown.labour_charge_to_client or 0)
                result.materials_charge = float(breakdown.materials_parking_cc_charge or 0)
                result.final_total = float(breakdown.final_total or 0)
                result.profit_gbp = float(breakdown.profit_gbp or 0)
                result.profit_pct = float(breakdown.profit_pct or 0)
                result.formula_source = breakdown.formula_source
                result.internal_notes = breakdown.internal_notes or ""

                if ws is not None:
                    ws.cell(row=row_num, column=8, value=result.labour_charge)
                    ws.cell(row=row_num, column=9, value=result.materials_charge)
                    ws.cell(row=row_num, column=10, value=result.final_total)
                    ws.cell(row=row_num, column=11, value=result.profit_gbp)
                    ws.cell(row=row_num, column=12, value=result.profit_pct)
                    ws.cell(row=row_num, column=13, value=result.formula_source)
                    ws.cell(row=row_num, column=14, value=result.status)
                    ws.cell(row=row_num, column=15, value=None)
                    notes_cell = ws.cell(row=row_num, column=16, value=result.internal_notes)
                    notes_cell.alignment = wrap
            except (ValueError, InvalidOperation) as exc:
                result.error = str(exc)
                if ws is not None:
                    ws.cell(row=row_num, column=14, value="ERROR")
                    ws.cell(row=row_num, column=15, value=result.error)
            except Exception as exc:
                result.error = f"{type(exc).__name__}: {exc}"
                if ws is not None:
                    ws.cell(row=row_num, column=14, value="ERROR")
                    ws.cell(row=row_num, column=15, value=result.error)

            results.append(result)
            processed += 1

            if processed % 100 == 0:
                print(f"Processed {processed} rows...")
    finally:
        db.close()

    if output_path is not None and wb_out is not None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb_out.save(output_path)
    if wb_out is not None:
        wb_out.close()
    wb_data.close()
    return results


def write_report(
    results: list[RowResult],
    report_path: Path,
    *,
    input_path: Path,
    output_path: Path | None,
    csv_path: Path | None = None,
) -> None:
    ok = sum(1 for r in results if r.status == "OK")
    errors = [r for r in results if r.status != "OK"]
    lines = [
        "# Book2 Batch Calculation Report",
        "",
        f"Input: `{input_path}`",
        f"Output: `{output_path or csv_path}`",
        "",
        "## Summary",
        "",
        f"- Rows processed: **{len(results)}**",
        f"- OK: **{ok}**",
        f"- ERROR: **{len(errors)}**",
        "",
    ]

    if results:
        lines.extend(["## Sample results (first 3 OK rows)", ""])
        shown = 0
        for item in results:
            if item.status != "OK":
                continue
            lines.append(
                f"- Row {item.row_num}: {item.client} / {item.trade} — "
                f"Labour £{item.labour_charge:.2f}, Mat/CC £{item.materials_charge:.2f}, "
                f"Total £{item.final_total:.2f}, Profit £{item.profit_gbp:.2f} ({item.profit_pct:.1f}%)"
            )
            shown += 1
            if shown >= 3:
                break

    if errors:
        lines.extend(["", "## Errors", ""])
        for item in errors[:20]:
            lines.append(f"- Row {item.row_num} ({item.client} / {item.trade}): {item.error}")
        if len(errors) > 20:
            lines.append(f"- ... and {len(errors) - 20} more")

    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=None, help="Write results to XLSX (default when --csv-output not set)")
    parser.add_argument("--csv-output", type=Path, default=None, help="Write results to CSV")
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--limit", type=int, default=None, help="Process only the first N data rows")
    parser.add_argument("--no-report", action="store_true")
    args = parser.parse_args()

    if not args.input.exists():
        raise FileNotFoundError(f"Input workbook not found: {args.input}")

    output_path = args.output
    if output_path is None and args.csv_output is None:
        output_path = DEFAULT_OUTPUT

    print(f"Reading: {args.input}")
    results = process_workbook(args.input, output_path, limit=args.limit)
    ok = sum(1 for r in results if r.status == "OK")

    if output_path is not None:
        print(f"Wrote: {output_path}")
    if args.csv_output is not None:
        write_results_csv(results, args.csv_output)
        print(f"Wrote CSV: {args.csv_output}")

    print(f"Processed {len(results)} rows — OK: {ok}, ERROR: {len(results) - ok}")

    if not args.no_report:
        write_report(
            results,
            args.report,
            input_path=args.input,
            output_path=output_path,
            csv_path=args.csv_output,
        )
        print(f"Report: {args.report}")

    if results and ok == 0:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
