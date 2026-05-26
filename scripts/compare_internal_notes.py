#!/usr/bin/env python3
"""Compare generated internal notes against Excel golden file."""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from decimal import InvalidOperation
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
SCRIPTS = ROOT / "scripts"

for path in (BACKEND, SCRIPTS, BACKEND / "tests"):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from helpers.corner_cases_workbook import FIXTURE_PATH, load_corner_case_rows
from helpers.internal_notes import normalize_internal_notes_for_comparison
from run_book2_batch_calculations import _calculate_row, _setup_backend_path


@dataclass
class CompareResult:
    row_num: int
    client: str
    trade: str
    expected: str
    actual: str
    match: bool


def compare_workbook(
    input_path: Path | None = None,
    *,
    limit: int | None = None,
    write_back: bool = False,
) -> tuple[list[CompareResult], Path | None]:
    import openpyxl

    _setup_backend_path()
    from app.db.session import SessionLocal

    path = input_path or FIXTURE_PATH
    rows = load_corner_case_rows(path, limit=limit)

    actual_col = None
    wb_out = None
    ws_out = None
    if write_back and path.exists():
        wb_out = openpyxl.load_workbook(path)
        ws_out = wb_out.active
        for col in range(1, ws_out.max_column + 1):
            if ws_out.cell(row=1, column=col).value == "internal_notes_my code":
                actual_col = col
                break

    results: list[CompareResult] = []
    db = SessionLocal()

    try:
        for index, row in enumerate(rows):
            try:
                breakdown = _calculate_row(
                    db,
                    row.trade,
                    row.client,
                    row.engineers,
                    row.hours,
                    row.parking,
                    row.congestion,
                    row.materials,
                )
                actual = breakdown.internal_notes or ""
            except (ValueError, InvalidOperation) as exc:
                actual = f"ERROR: {exc}"

            expected = row.expected_internal_notes
            match = expected == actual
            results.append(
                CompareResult(
                    row_num=row.row_num,
                    client=row.client,
                    trade=row.trade,
                    expected=expected,
                    actual=actual,
                    match=match,
                )
            )
            if ws_out is not None and actual_col is not None:
                ws_out.cell(row=row.row_num, column=actual_col, value=actual)

            if (index + 1) % 100 == 0:
                print(f"Compared {index + 1} rows...")
    finally:
        db.close()

    output_path = None
    if wb_out is not None:
        wb_out.save(path)
        wb_out.close()
        output_path = path

    return results, output_path


def print_summary(results: list[CompareResult]) -> None:
    matched = sum(1 for item in results if item.match)
    mismatches = [item for item in results if not item.match]
    total = len(results)
    print(f"Exact match: {matched}/{total}")
    print(f"Mismatches: {len(mismatches)}")
    if mismatches:
        print("\nFirst 10 mismatches:")
        for item in mismatches[:10]:
            print(f"\n--- Row {item.row_num} ({item.client} / {item.trade}) ---")
            exp_lines = normalize_internal_notes_for_comparison(item.expected).splitlines()
            act_lines = normalize_internal_notes_for_comparison(item.actual).splitlines()
            for idx, (exp_line, act_line) in enumerate(zip(exp_lines, act_lines), start=1):
                if exp_line != act_line:
                    print(f"  L{idx} expected: {exp_line!r}")
                    print(f"  L{idx} actual:   {act_line!r}")
            if len(exp_lines) != len(act_lines):
                print(f"  Line count expected={len(exp_lines)} actual={len(act_lines)}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=None, help=f"Defaults to {FIXTURE_PATH}")
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--write-back", action="store_true", help="Update internal_notes_my code column")
    args = parser.parse_args()

    input_path = args.input or FIXTURE_PATH
    if not input_path.exists():
        raise FileNotFoundError(f"Workbook not found: {input_path}")

    print(f"Comparing: {input_path}")
    results, output_path = compare_workbook(input_path, limit=args.limit, write_back=args.write_back)
    if output_path is not None:
        print(f"Updated: {output_path}")
    print_summary(results)
    if results and not all(item.match for item in results):
        raise SystemExit(1)


if __name__ == "__main__":
    main()
