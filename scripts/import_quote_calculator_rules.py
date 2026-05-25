#!/usr/bin/env python3
"""Import clients, trades, and XLSX rate rules from docs/1.7 MASTER HELPER.xlsx."""

from __future__ import annotations

import argparse
import getpass
import os
import sys
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path

import openpyxl
from sqlalchemy import select, update

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
sys.path.insert(0, str(BACKEND))

from app.db.session import SessionLocal
from app.models.rate_rule import RateRule
from app.services.client_service import find_client_by_name_or_alias, get_or_create_client_for_import
from app.services.trade_service import get_or_create_trade_for_import
from app.models.trade import Trade


XLSX_PATH = ROOT / "docs" / "1.7 MASTER HELPER.xlsx"
RULE_VERSION = "xlsx-master-helper-1.7"
ACTIVE_FROM = date(2024, 1, 1)
EXPECTED_FULL_IMPORT_CLIENTS = 158
EXPECTED_FULL_IMPORT_TRADES = 16
EXPECTED_FULL_IMPORT_RULES = 2528

XLSX_DEFAULTS = {
    "hourly_overhead_pct": Decimal("0.30"),
    "daily_overhead_pct": Decimal("0.20"),
    "daily_overhead_long_job_pct": Decimal("0.15"),
    "labourer_hourly_cost": Decimal("18.75"),
    "labourer_daily_cost": Decimal("150"),
    "material_charge_denominator": Decimal("0.20"),
    "parking_charge_denominator": Decimal("0.20"),
    "congestion_charge_denominator": Decimal("0.20"),
    "mround_increment": Decimal("5"),
    "oj_uplift_pct": Decimal("10"),
    "nhs_overhead_uplift_pct": Decimal("15"),
    "eaf_flat_fee": Decimal("1"),
}


def _empty_stats(*, clients: int = 0, trades: int = 0) -> dict[str, int]:
    return {
        "clients": clients,
        "trades": trades,
        "clients_created": 0,
        "clients_updated": 0,
        "trades_created": 0,
        "trades_updated": 0,
        "rules_created": 0,
        "rules_updated": 0,
        "rules_deactivated": 0,
        "rules_skipped": 0,
        "rules_would_create": clients * trades,
    }


def _dec(value) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        cleaned = value.strip().replace("%", "")
        if not cleaned or any(ch.isalpha() for ch in cleaned):
            return None
        if "£" in cleaned:
            cleaned = cleaned.split("£")[-1].strip()
        try:
            return Decimal(cleaned)
        except Exception:
            return None
    return Decimal(str(value))


def _normalize_fee(value) -> Decimal:
    fee = _dec(value) or Decimal("0")
    if fee > Decimal("1"):
        fee = fee / Decimal("100")
    return fee


def _normalize_name(name: str) -> str:
    return " ".join(name.strip().split())


def load_trade_rates(ws) -> list[dict]:
    trades: list[dict] = []
    for row in range(3, 19):
        trade = ws[f"AN{row}"].value
        hourly_client = ws[f"AO{row}"].value
        subby_daily = ws[f"AR{row}"].value
        direct_hourly = ws[f"AT{row}"].value
        if not trade or not hourly_client:
            continue
        daily = _dec(subby_daily)
        hourly_direct = _dec(direct_hourly)
        trades.append(
            {
                "name": _normalize_name(str(trade)),
                "hourly_client_rate": _dec(hourly_client),
                "day_rate": daily,
                "direct_daily_cost": daily,
                "direct_hourly_cost": hourly_direct,
            }
        )
    return trades


def load_client_fees(ws_quote, ws_sla) -> dict[str, Decimal]:
    fees: dict[str, Decimal] = {}
    for row in range(2, 120):
        name = ws_quote[f"AY{row}"].value
        fee = ws_quote[f"AZ{row}"].value
        if name and str(name).strip() and fee is not None and str(name) != "SELECT AGENT:":
            fees[_normalize_name(str(name))] = _normalize_fee(fee)

    for row in range(5, ws_sla.max_row + 1):
        name = ws_sla.cell(row, 1).value
        fee = ws_sla.cell(row, 6).value
        if not name:
            continue
        key = _normalize_name(str(name))
        if key not in fees and fee is not None:
            if isinstance(fee, str) and "%" in fee:
                continue
            fees[key] = _normalize_fee(fee)
    return fees


def _rule_payload(*, client_id, trade_id, client_name: str, trade: dict, client_fee_pct: Decimal) -> dict:
    day_rate = trade["day_rate"]
    return {
        "client_id": client_id,
        "trade_id": trade_id,
        "version": RULE_VERSION,
        "hourly_rate": trade["hourly_client_rate"],
        "half_day_rate": round(day_rate / Decimal("2"), 2) if day_rate else None,
        "day_rate": day_rate,
        "direct_hourly_cost": trade["direct_hourly_cost"],
        "direct_daily_cost": trade["direct_daily_cost"],
        "material_markup_type": "percentage",
        "material_markup_value": Decimal("20"),
        "vat_rate": Decimal("20"),
        "approval_threshold": Decimal("5000"),
        "minimum_margin_percentage": Decimal("10"),
        "active_from": ACTIVE_FROM,
        "is_active": True,
        "client_fee_pct": client_fee_pct,
        "formula_source": "xlsx",
        "xlsx_client_name": client_name,
        "xlsx_trade_name": trade["name"],
        **XLSX_DEFAULTS,
    }


def upsert_rate_rule(
    db,
    *,
    client_id,
    trade_id,
    client_name: str,
    trade: dict,
    client_fee_pct: Decimal,
    dry_run: bool,
    overwrite: bool,
) -> str:
    existing = db.scalar(
        select(RateRule).where(
            RateRule.client_id == client_id,
            RateRule.trade_id == trade_id,
            RateRule.version == RULE_VERSION,
        )
    )
    if not existing and overwrite:
        existing = db.scalar(
            select(RateRule).where(
                RateRule.client_id == client_id,
                RateRule.trade_id == trade_id,
            ).order_by(RateRule.created_at.desc())
        )

    payload = _rule_payload(
        client_id=client_id,
        trade_id=trade_id,
        client_name=client_name,
        trade=trade,
        client_fee_pct=client_fee_pct,
    )

    if existing:
        if not overwrite:
            return "skipped" if not dry_run else "would_skip"
        if dry_run:
            return "would_update"
        for key, value in payload.items():
            setattr(existing, key, value)
        return "updated"

    if dry_run:
        return "would_create"

    db.add(RateRule(**payload))
    return "created"


def _is_full_import(client_filter: str | None, trade_filter: str | None) -> bool:
    return client_filter is None and trade_filter is None


def _requires_destructive_confirmation(*, overwrite: bool, deactivate_existing: bool, full_import: bool) -> bool:
    return deactivate_existing or (overwrite and full_import)


def _confirm_destructive_import(*, overwrite: bool, deactivate_existing: bool, full_import: bool) -> None:
    actions: list[str] = []
    if deactivate_existing:
        actions.append("deactivate ALL existing rate_rules")
    if overwrite and full_import:
        actions.append("overwrite existing client/trade rules with XLSX data")
    warning = (
        "DESTRUCTIVE XLSX IMPORT\n"
        f"- Workbook: {XLSX_PATH}\n"
        f"- Expected XLSX rules: {EXPECTED_FULL_IMPORT_RULES}\n"
        f"- Actions: {', '.join(actions)}\n"
        "Ensure you have exported rate_rules (scripts/export_rate_rules.py or pg_dump) before continuing.\n"
        "Type YES to proceed: "
    )
    response = input(warning).strip()
    if response != "YES":
        raise SystemExit("Import aborted: destructive confirmation not provided.")


def _log_import_audit(db, *, action: str, stats: dict, initiated_by: str | None, error: str | None = None) -> None:
    _ = (db, action, stats, initiated_by, error)


def _format_summary(stats: dict[str, int], *, dry_run: bool) -> str:
    mode = "DRY RUN" if dry_run else "IMPORT"
    return (
        f"[{mode}] clients={stats['clients']} trades={stats['trades']} "
        f"clients_created={stats['clients_created']} clients_updated={stats['clients_updated']} "
        f"trades_created={stats['trades_created']} trades_updated={stats['trades_updated']} "
        f"rules_created={stats['rules_created']} rules_updated={stats['rules_updated']} "
        f"rules_deactivated={stats['rules_deactivated']} rules_skipped={stats['rules_skipped']} "
        f"rules_would_create={stats['rules_would_create']}"
    )


def import_rules(
    *,
    dry_run: bool = False,
    client_filter: str | None = None,
    trade_filter: str | None = None,
    overwrite: bool = False,
    deactivate_existing: bool = False,
    initiated_by: str | None = None,
    db=None,
) -> dict[str, int]:
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Missing workbook: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws_quote = wb["QUOTE CALCULATOR"]
    ws_sla = wb["SLA TABLE"]
    trades = load_trade_rates(ws_quote)
    client_fees = load_client_fees(ws_quote, ws_sla)

    if client_filter:
        client_fees = {k: v for k, v in client_fees.items() if client_filter.lower() in k.lower()}
    if trade_filter:
        trades = [t for t in trades if trade_filter.lower() in t["name"].lower()]

    stats = _empty_stats(clients=len(client_fees), trades=len(trades))

    own_session = db is None
    if not dry_run and own_session:
        db = SessionLocal()

    try:
        if dry_run:
            if db is None:
                stats["clients_created"] = len(client_fees)
                stats["trades_created"] = len(trades)
                stats["rules_would_create"] = len(client_fees) * len(trades)
                return stats

            for trade in trades:
                existing = db.scalar(select(Trade).where(Trade.name == trade["name"]))
                if existing:
                    stats["trades_updated"] += 1
                else:
                    stats["trades_created"] += 1

            for source_client_name in sorted(client_fees):
                existing = find_client_by_name_or_alias(db, source_client_name)
                if existing:
                    stats["clients_updated"] += 1
                else:
                    stats["clients_created"] += 1

                client = existing
                if client is None:
                    for trade in trades:
                        stats["rules_would_create"] += 1
                    continue

                for trade in trades:
                    trade_model = db.scalar(select(Trade).where(Trade.name == trade["name"]))
                    if trade_model is None:
                        stats["rules_would_create"] += 1
                        continue
                    action = upsert_rate_rule(
                        db,
                        client_id=client.id,
                        trade_id=trade_model.id,
                        client_name=source_client_name,
                        trade=trade,
                        client_fee_pct=client_fees[source_client_name],
                        dry_run=True,
                        overwrite=overwrite,
                    )
                    if action == "would_create":
                        stats["rules_would_create"] += 1
                    elif action == "would_update":
                        stats["rules_updated"] += 1
                    elif action == "would_skip":
                        stats["rules_skipped"] += 1
            return stats

        stats["rules_would_create"] = 0
        _log_import_audit(
            db,
            action="xlsx_import_started",
            stats=stats,
            initiated_by=initiated_by,
        )
        if deactivate_existing:
            result = db.execute(update(RateRule).values(is_active=False))
            stats["rules_deactivated"] = result.rowcount or 0

        trade_models: dict[str, object] = {}
        for trade in trades:
            trade_model, created = get_or_create_trade_for_import(db, trade["name"], source_label=XLSX_PATH.name)
            trade_models[trade["name"]] = trade_model
            if created:
                stats["trades_created"] += 1
            else:
                stats["trades_updated"] += 1

        for source_client_name in sorted(client_fees):
            client, created, alias_added = get_or_create_client_for_import(db, source_client_name)
            if created:
                stats["clients_created"] += 1
            elif alias_added:
                stats["clients_updated"] += 1

            for trade in trades:
                action = upsert_rate_rule(
                    db,
                    client_id=client.id,
                    trade_id=trade_models[trade["name"]].id,
                    client_name=source_client_name,
                    trade=trade,
                    client_fee_pct=client_fees[source_client_name],
                    dry_run=False,
                    overwrite=overwrite,
                )
                if action == "created":
                    stats["rules_created"] += 1
                elif action == "updated":
                    stats["rules_updated"] += 1
                elif action == "skipped":
                    stats["rules_skipped"] += 1

        _log_import_audit(
            db,
            action="xlsx_import_completed",
            stats=stats,
            initiated_by=initiated_by,
        )
        if own_session:
            db.commit()
    except Exception as exc:
        if own_session and db is not None:
            db.rollback()
        if db is not None:
            try:
                _log_import_audit(
                    db,
                    action="xlsx_import_failed",
                    stats=stats,
                    initiated_by=initiated_by,
                    error=str(exc),
                )
                if own_session:
                    db.commit()
            except Exception:
                if own_session:
                    db.rollback()
        raise
    finally:
        if own_session and db is not None:
            db.close()

    return stats


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true", help="Parse workbook and report counts without writing")
    parser.add_argument("--client", help="Only import rules for clients matching this substring")
    parser.add_argument("--trade", help="Only import rules for trades matching this substring")
    parser.add_argument("--overwrite", action="store_true", help="Update existing client/trade rules if version differs")
    parser.add_argument(
        "--deactivate-existing",
        action="store_true",
        help="Deactivate all existing rate rules before import",
    )
    parser.add_argument(
        "--confirm-destructive",
        action="store_true",
        help="Required with --deactivate-existing or full --overwrite import; skips interactive prompt",
    )
    args = parser.parse_args()

    full_import = _is_full_import(args.client, args.trade)
    destructive = _requires_destructive_confirmation(
        overwrite=args.overwrite,
        deactivate_existing=args.deactivate_existing,
        full_import=full_import,
    )
    if destructive and not args.dry_run:
        if args.confirm_destructive:
            print(
                "[WARN] Destructive import confirmed via --confirm-destructive. "
                "Ensure rate_rules backup exists."
            )
        else:
            _confirm_destructive_import(
                overwrite=args.overwrite,
                deactivate_existing=args.deactivate_existing,
                full_import=full_import,
            )

    initiated_by = os.environ.get("IMPORT_INITIATED_BY") or getpass.getuser()

    stats = import_rules(
        dry_run=args.dry_run,
        client_filter=args.client,
        trade_filter=args.trade,
        overwrite=args.overwrite,
        deactivate_existing=args.deactivate_existing,
        initiated_by=initiated_by,
    )
    print(_format_summary(stats, dry_run=args.dry_run))


if __name__ == "__main__":
    main()
