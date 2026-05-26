#!/usr/bin/env python3
"""Regenerate corner_cases_days_internal_notes_all.xlsx expected-output columns
using the current calculation engine.

Run inside the backend Docker container:
    python /workspace/scripts/regenerate_days_fixture.py
"""

from __future__ import annotations

import sys
from decimal import Decimal
from pathlib import Path

ROOT = Path("/workspace")
BACKEND = ROOT / "backend"
SCRIPTS = ROOT / "scripts"

for p in (BACKEND, SCRIPTS):
    if str(p) not in sys.path:
        sys.path.insert(0, str(p))

# Columns in the daily fixture (1-based)
DAYS_COL_TRADE = 2   # B
DAYS_COL_CLIENT = 3  # C
DAYS_COL_ENGINEERS = 4  # D
DAYS_COL_DAYS = 5    # E
DAYS_COL_PARKING = 6  # F
DAYS_COL_CONGESTION = 7  # G
DAYS_COL_MATERIALS = 8  # H

# Expected-output columns (written back)
COL_LABOUR_CHARGE = 14   # N
COL_MATERIALS_CHARGE = 15  # O
COL_TOTAL_CHARGE = 16    # P
COL_PROFIT_GBP = 20      # T
COL_PROFIT_PCT = 21      # U
DAYS_COL_INTERNAL_NOTES = 30  # AD

FIXTURE = ROOT / "tests" / "fixtures" / "corner_cases_days_internal_notes_all.xlsx"
OUTPUT = Path("/tmp/corner_cases_days_internal_notes_all.xlsx")


def _to_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace("£", "")
    return Decimal(text) if text else Decimal("0")


def main() -> None:
    import openpyxl

    from app.core.config import settings
    from app.db.session import SessionLocal
    from app.engines.approval_engine import build_calculation_breakdown
    from app.engines.rules_engine import find_active_rule
    from app.engines.xlsx_breakdown_engine import config_from_rule, trade_from_rule
    from app.engines.xlsx_quote_calculator import (
        calculate_daily_quote,
        format_notes_job_pct,
        format_notes_profit,
    )
    from app.models.trade import Trade
    from app.schemas.calculation import ChargeInput, LabourInput, MaterialInput
    from app.services.client_service import find_client_by_name_or_alias
    from sqlalchemy import select

    print(f"Loading fixture: {FIXTURE}")
    wb = openpyxl.load_workbook(FIXTURE)
    ws = wb.active

    db = SessionLocal()
    ok = errors = 0

    try:
        for row_num in range(2, ws.max_row + 1):
            trade_val = ws.cell(row=row_num, column=DAYS_COL_TRADE).value
            if trade_val is None or str(trade_val).strip() == "":
                continue

            trade_name = str(trade_val).strip()
            client_name = str(ws.cell(row=row_num, column=DAYS_COL_CLIENT).value or "").strip()
            engineers = _to_decimal(ws.cell(row=row_num, column=DAYS_COL_ENGINEERS).value)
            days = _to_decimal(ws.cell(row=row_num, column=DAYS_COL_DAYS).value)
            parking = _to_decimal(ws.cell(row=row_num, column=DAYS_COL_PARKING).value)
            congestion = _to_decimal(ws.cell(row=row_num, column=DAYS_COL_CONGESTION).value)
            materials = _to_decimal(ws.cell(row=row_num, column=DAYS_COL_MATERIALS).value)

            try:
                client = find_client_by_name_or_alias(db, client_name)
                if client is None:
                    raise ValueError(f"Client not found: {client_name}")

                trade_obj = db.scalar(select(Trade).where(Trade.name == trade_name))
                if trade_obj is None:
                    raise ValueError(f"Trade not found: {trade_name}")

                matched = find_active_rule(db, client.id, trade_obj.id, None)
                if matched is None:
                    raise ValueError(f"No active rule for {client_name}/{trade_name}")

                material_items = []
                if materials > 0:
                    material_items.append(
                        MaterialInput(material_name="Materials", quantity=Decimal("1"), unit_cost=materials)
                    )

                charges = None
                if parking > 0 or congestion > 0:
                    charges = ChargeInput(
                        parking_required=parking > 0,
                        parking_type="fixed" if parking > 0 else None,
                        parking_fixed_amount=parking if parking > 0 else None,
                        congestion_required=congestion > 0,
                        congestion_amount=congestion,
                    )

                breakdown = build_calculation_breakdown(
                    labour_items=[
                        LabourInput(
                            labour_type="day",
                            number_of_engineers=int(engineers),
                            days_on_site=days,
                            number_of_labourers=0,
                        )
                    ],
                    material_items=material_items,
                    charges=charges,
                    matched_rule=matched,
                    formula_version=settings.formula_version,
                )

                rule = matched.rule
                rates = trade_from_rule(rule)
                config = config_from_rule(rule)
                quote_result = calculate_daily_quote(
                    trade=rates,
                    engineers=int(engineers),
                    days=days,
                    labourers=0,
                    client_fee_pct=rule.client_fee_pct,
                    parking=parking,
                    congestion=congestion,
                    materials=materials,
                    config=config,
                )

                labour_charge = breakdown.labour_charge_to_client or Decimal("0")
                materials_charge = breakdown.materials_parking_cc_charge or Decimal("0")
                total_charge = labour_charge + materials_charge
                profit_gbp = int(format_notes_profit(quote_result))
                profit_pct = format_notes_job_pct(quote_result.profit_gbp, quote_result.profit_pct, total_charge)

                ws.cell(row=row_num, column=COL_LABOUR_CHARGE).value = float(labour_charge)
                ws.cell(row=row_num, column=COL_MATERIALS_CHARGE).value = float(materials_charge)
                ws.cell(row=row_num, column=COL_TOTAL_CHARGE).value = float(total_charge)
                ws.cell(row=row_num, column=COL_PROFIT_GBP).value = profit_gbp
                ws.cell(row=row_num, column=COL_PROFIT_PCT).value = profit_pct
                ws.cell(row=row_num, column=DAYS_COL_INTERNAL_NOTES).value = breakdown.internal_notes or ""

                ok += 1
                if ok % 100 == 0:
                    print(f"  Processed {ok} rows OK so far...")

            except Exception as exc:
                errors += 1
                print(f"  Row {row_num} ({client_name}/{trade_name}) ERROR: {exc}")

    finally:
        db.close()

    wb.save(OUTPUT)
    wb.close()
    print(f"\nDone. OK={ok}, Errors={errors}")
    print(f"Saved: {OUTPUT}")


if __name__ == "__main__":
    main()
