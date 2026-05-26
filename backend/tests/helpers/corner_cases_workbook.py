"""Load corner-case golden rows from the committed Excel fixture."""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
FIXTURE_PATH = ROOT / "tests" / "fixtures" / "corner_cases_internal_notes_all.xlsx"
DAYS_FIXTURE_PATH = ROOT / "tests" / "fixtures" / "corner_cases_days_internal_notes_all.xlsx"

# Hourly input columns
COL_TRADE = 2  # B
COL_CLIENT = 3  # C
COL_ENGINEERS = 4  # D
COL_HOURS = 5  # E
COL_PARKING = 7  # G
COL_CONGESTION = 8  # H
COL_MATERIALS = 9  # I

# Days input columns
DAYS_COL_TRADE = 2  # B
DAYS_COL_CLIENT = 3  # C
DAYS_COL_ENGINEERS = 4  # D
DAYS_COL_DAYS = 5  # E
DAYS_COL_PARKING = 6  # F
DAYS_COL_CONGESTION = 7  # G
DAYS_COL_MATERIALS = 8  # H

# Expected output columns (shared N-P, T-U)
COL_LABOUR_CHARGE = 14  # N
COL_MATERIALS_CHARGE = 15  # O
COL_TOTAL_CHARGE = 16  # P
COL_PROFIT_GBP = 20  # T
COL_PROFIT_PCT = 21  # U
COL_INTERNAL_NOTES = 31  # AE (hourly)
DAYS_COL_INTERNAL_NOTES = 30  # AD (days)


@dataclass(frozen=True)
class CornerCaseRow:
    row_num: int
    trade: str
    client: str
    engineers: Decimal
    hours: Decimal
    parking: Decimal
    congestion: Decimal
    materials: Decimal
    expected_labour_charge: Decimal
    expected_materials_charge: Decimal
    expected_total_charge: Decimal
    expected_profit_gbp: int
    expected_profit_pct: int
    expected_internal_notes: str


@dataclass(frozen=True)
class CornerCaseDaysRow:
    row_num: int
    trade: str
    client: str
    engineers: Decimal
    days: Decimal
    parking: Decimal
    congestion: Decimal
    materials: Decimal
    expected_labour_charge: Decimal
    expected_materials_charge: Decimal
    expected_total_charge: Decimal
    expected_profit_gbp: int
    expected_profit_pct: int
    expected_internal_notes: str


GoldenCaseRow = CornerCaseRow | CornerCaseDaysRow


def _to_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace("£", "")
    if not text:
        return Decimal("0")
    return Decimal(text)


def _to_int(value) -> int:
    if value is None or value == "":
        return 0
    return int(Decimal(str(value)))


def load_corner_case_rows(
    workbook_path: Path | None = None,
    *,
    limit: int | None = None,
) -> list[CornerCaseRow]:
    import openpyxl

    path = workbook_path or FIXTURE_PATH
    if not path.exists():
        raise FileNotFoundError(f"Corner cases workbook not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows: list[CornerCaseRow] = []

    try:
        for row_num in range(2, ws.max_row + 1):
            trade = ws.cell(row=row_num, column=COL_TRADE).value
            if trade is None or str(trade).strip() == "":
                continue
            if limit is not None and len(rows) >= limit:
                break

            notes = ws.cell(row=row_num, column=COL_INTERNAL_NOTES).value
            rows.append(
                CornerCaseRow(
                    row_num=row_num,
                    trade=str(trade).strip(),
                    client=str(ws.cell(row=row_num, column=COL_CLIENT).value or "").strip(),
                    engineers=_to_decimal(ws.cell(row=row_num, column=COL_ENGINEERS).value),
                    hours=_to_decimal(ws.cell(row=row_num, column=COL_HOURS).value),
                    parking=_to_decimal(ws.cell(row=row_num, column=COL_PARKING).value),
                    congestion=_to_decimal(ws.cell(row=row_num, column=COL_CONGESTION).value),
                    materials=_to_decimal(ws.cell(row=row_num, column=COL_MATERIALS).value),
                    expected_labour_charge=_to_decimal(ws.cell(row=row_num, column=COL_LABOUR_CHARGE).value),
                    expected_materials_charge=_to_decimal(ws.cell(row=row_num, column=COL_MATERIALS_CHARGE).value),
                    expected_total_charge=_to_decimal(ws.cell(row=row_num, column=COL_TOTAL_CHARGE).value),
                    expected_profit_gbp=_to_int(ws.cell(row=row_num, column=COL_PROFIT_GBP).value),
                    expected_profit_pct=_to_int(ws.cell(row=row_num, column=COL_PROFIT_PCT).value),
                    expected_internal_notes=str(notes or ""),
                )
            )
    finally:
        wb.close()

    return rows


def load_corner_case_days_rows(
    workbook_path: Path | None = None,
    *,
    limit: int | None = None,
) -> list[CornerCaseDaysRow]:
    import openpyxl

    path = workbook_path or DAYS_FIXTURE_PATH
    if not path.exists():
        raise FileNotFoundError(f"Corner cases days workbook not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows: list[CornerCaseDaysRow] = []

    try:
        for row_num in range(2, ws.max_row + 1):
            trade = ws.cell(row=row_num, column=DAYS_COL_TRADE).value
            if trade is None or str(trade).strip() == "":
                continue
            if limit is not None and len(rows) >= limit:
                break

            notes = ws.cell(row=row_num, column=DAYS_COL_INTERNAL_NOTES).value
            rows.append(
                CornerCaseDaysRow(
                    row_num=row_num,
                    trade=str(trade).strip(),
                    client=str(ws.cell(row=row_num, column=DAYS_COL_CLIENT).value or "").strip(),
                    engineers=_to_decimal(ws.cell(row=row_num, column=DAYS_COL_ENGINEERS).value),
                    days=_to_decimal(ws.cell(row=row_num, column=DAYS_COL_DAYS).value),
                    parking=_to_decimal(ws.cell(row=row_num, column=DAYS_COL_PARKING).value),
                    congestion=_to_decimal(ws.cell(row=row_num, column=DAYS_COL_CONGESTION).value),
                    materials=_to_decimal(ws.cell(row=row_num, column=DAYS_COL_MATERIALS).value),
                    expected_labour_charge=_to_decimal(ws.cell(row=row_num, column=COL_LABOUR_CHARGE).value),
                    expected_materials_charge=_to_decimal(ws.cell(row=row_num, column=COL_MATERIALS_CHARGE).value),
                    expected_total_charge=_to_decimal(ws.cell(row=row_num, column=COL_TOTAL_CHARGE).value),
                    expected_profit_gbp=_to_int(ws.cell(row=row_num, column=COL_PROFIT_GBP).value),
                    expected_profit_pct=_to_int(ws.cell(row=row_num, column=COL_PROFIT_PCT).value),
                    expected_internal_notes=str(notes or ""),
                )
            )
    finally:
        wb.close()

    return rows


@dataclass
class GoldenCompareMismatch:
    row_num: int
    client: str
    trade: str
    field: str
    expected: object
    actual: object


def calculate_hourly_result(db, row: CornerCaseRow):
    """Return HourlyQuoteResult for display-profit helpers (columns T/U)."""
    from sqlalchemy import select

    from app.engines.rules_engine import find_active_rule
    from app.engines.xlsx_breakdown_engine import config_from_rule, trade_from_rule
    from app.engines.xlsx_quote_calculator import calculate_hourly_quote
    from app.models.trade import Trade
    from app.services.client_service import find_client_by_name_or_alias

    client = find_client_by_name_or_alias(db, row.client)
    if client is None:
        raise ValueError(f"Client not found: {row.client}")

    trade = db.scalar(select(Trade).where(Trade.name == row.trade.strip()))
    if trade is None:
        raise ValueError(f"Trade not found: {row.trade}")

    matched = find_active_rule(db, client.id, trade.id, None)
    if matched is None:
        raise ValueError(f"No active rate rule for client '{row.client}' / trade '{row.trade}'")

    rule = matched.rule
    rates = trade_from_rule(rule)
    config = config_from_rule(rule)
    return calculate_hourly_quote(
        trade=rates,
        engineers=int(row.engineers),
        hours=row.hours,
        client_fee_pct=rule.client_fee_pct,
        parking=row.parking,
        congestion=row.congestion,
        materials=row.materials,
        config=config,
    )


def calculate_daily_result(db, row: CornerCaseDaysRow):
    """Return DailyQuoteResult for display-profit helpers (columns T/U)."""
    from sqlalchemy import select

    from app.engines.rules_engine import find_active_rule
    from app.engines.xlsx_breakdown_engine import config_from_rule, trade_from_rule
    from app.engines.xlsx_quote_calculator import calculate_daily_quote
    from app.models.trade import Trade
    from app.services.client_service import find_client_by_name_or_alias

    client = find_client_by_name_or_alias(db, row.client)
    if client is None:
        raise ValueError(f"Client not found: {row.client}")

    trade = db.scalar(select(Trade).where(Trade.name == row.trade.strip()))
    if trade is None:
        raise ValueError(f"Trade not found: {row.trade}")

    matched = find_active_rule(db, client.id, trade.id, None)
    if matched is None:
        raise ValueError(f"No active rate rule for client '{row.client}' / trade '{row.trade}'")

    rule = matched.rule
    rates = trade_from_rule(rule)
    config = config_from_rule(rule)
    return calculate_daily_quote(
        trade=rates,
        engineers=int(row.engineers),
        days=row.days,
        labourers=0,
        client_fee_pct=rule.client_fee_pct,
        parking=row.parking,
        congestion=row.congestion,
        materials=row.materials,
        config=config,
    )


def compare_row_to_breakdown(row: GoldenCaseRow, breakdown, quote_result) -> list[GoldenCompareMismatch]:
    from app.engines.xlsx_quote_calculator import format_notes_job_pct, format_notes_profit

    mismatches: list[GoldenCompareMismatch] = []

    labour = breakdown.labour_charge_to_client or Decimal("0")
    materials = breakdown.materials_parking_cc_charge or Decimal("0")
    total = labour + materials

    checks: list[tuple[str, object, object]] = [
        ("labour_charge", row.expected_labour_charge, labour),
        ("materials_charge", row.expected_materials_charge, materials),
        ("total_charge", row.expected_total_charge, total),
        (
            "profit_gbp_display",
            row.expected_profit_gbp,
            int(format_notes_profit(quote_result)),
        ),
        (
            "profit_pct_display",
            row.expected_profit_pct,
            format_notes_job_pct(quote_result.profit_gbp, quote_result.profit_pct, total),
        ),
        ("internal_notes", row.expected_internal_notes, breakdown.internal_notes or ""),
    ]

    for field, expected, actual in checks:
        if expected != actual:
            mismatches.append(
                GoldenCompareMismatch(
                    row_num=row.row_num,
                    client=row.client,
                    trade=row.trade,
                    field=field,
                    expected=expected,
                    actual=actual,
                )
            )

    return mismatches


def format_mismatch_report(mismatches: list[GoldenCompareMismatch], *, limit: int = 10) -> str:
    lines = [f"Mismatches: {len(mismatches)}"]
    for item in mismatches[:limit]:
        lines.append(
            f"  Row {item.row_num} ({item.client} / {item.trade}) "
            f"{item.field}: expected={item.expected!r} actual={item.actual!r}"
        )
    if len(mismatches) > limit:
        lines.append(f"  ... and {len(mismatches) - limit} more")
    return "\n".join(lines)
